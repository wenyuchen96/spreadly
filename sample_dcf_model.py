#!/usr/bin/env python3
"""
Create a sample DCF model XLSX file for testing upload
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

def create_sample_dcf_model():
    """Create a sample DCF model Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    
    # Header
    ws['A1'] = "TECHNOLOGY COMPANY DCF VALUATION MODEL"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Assumptions Section
    ws['A3'] = "KEY ASSUMPTIONS"
    ws['A3'].font = Font(bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    assumptions = [
        ["Revenue Growth Rate (Y1-3)", "25%"],
        ["Revenue Growth Rate (Y4-5)", "15%"],
        ["Terminal Growth Rate", "3%"],
        ["EBITDA Margin (Mature)", "30%"],
        ["Tax Rate", "25%"],
        ["WACC", "12%"],
        ["CapEx as % of Revenue", "3%"],
        ["Working Capital as % Rev", "5%"]
    ]
    
    for i, (label, value) in enumerate(assumptions, 4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    
    # Projection Years
    years = ["Year", "1", "2", "3", "4", "5"]
    for i, year in enumerate(years):
        cell = ws.cell(row=3, column=8+i, value=year)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Revenue Projections
    projections = [
        ["Revenue", 100000, "=I4*(1+$B$4)", "=J4*(1+$B$4)", "=K4*(1+$B$5)", "=L4*(1+$B$5)"],
        ["EBITDA", "=I4*0.20", "=J4*0.25", "=K4*$B$7", "=L4*$B$7", "=M4*$B$7"],
        ["Free Cash Flow", "=I5*0.8", "=J5*0.8", "=K5*0.8", "=L5*0.8", "=M5*0.8"]
    ]
    
    for i, row_data in enumerate(projections, 4):
        for j, value in enumerate(row_data):
            ws.cell(row=i, column=8+j, value=value)
    
    # Valuation Section
    ws['A15'] = "VALUATION SUMMARY"
    ws['A15'].font = Font(bold=True)
    ws['A15'].fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    valuation_data = [
        ["PV of FCF (Y1-5)", "=NPV($B$9,I6:M6)"],
        ["Terminal Value", "=M6*(1+$B$6)/($B$9-$B$6)/POWER(1+$B$9,5)"],
        ["Enterprise Value", "=B16+B17"],
        ["Equity Value", "=B18"]
    ]
    
    for i, (label, formula) in enumerate(valuation_data, 16):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = formula
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    # Save file
    output_path = Path("/Users/wenyuc/Dev/spreadly/sample_tech_dcf_model.xlsx")
    wb.save(output_path)
    print(f"✅ Created sample DCF model: {output_path}")
    return output_path

if __name__ == "__main__":
    create_sample_dcf_model()