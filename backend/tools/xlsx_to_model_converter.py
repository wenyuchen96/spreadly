#!/usr/bin/env python3
"""
Tool to convert XLSX financial models to RAG-compatible FinancialModel objects
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, List, Optional
from pathlib import Path
import json

from app.models.financial_model import (
    FinancialModel, ModelType, Industry, ComplexityLevel,
    ModelMetadata, PerformanceMetrics
)

class XLSXToModelConverter:
    """Convert XLSX financial models to FinancialModel objects for RAG"""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
    
    def convert_xlsx_to_model(
        self, 
        xlsx_path: str, 
        model_id: str,
        model_type: ModelType,
        industry: Industry = Industry.GENERAL,
        complexity: ComplexityLevel = ComplexityLevel.INTERMEDIATE
    ) -> FinancialModel:
        """
        Convert an XLSX file to a FinancialModel object
        
        Args:
            xlsx_path: Path to the XLSX file
            model_id: Unique identifier for the model
            model_type: Type of financial model (DCF, NPV, etc.)
            industry: Industry classification
            complexity: Complexity level
        """
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(xlsx_path, data_only=False)  # Keep formulas
        
        # Analyze the model structure
        analysis = self._analyze_excel_structure(workbook)
        
        # Generate Excel.js code from the structure
        excel_js_code = self._generate_excel_js_code(workbook, analysis)
        
        # Extract metadata
        metadata = self._extract_metadata(workbook, analysis)
        
        # Create the model
        model = FinancialModel(
            id=model_id,
            name=analysis['suggested_name'],
            description=analysis['description'],
            model_type=model_type,
            industry=industry,
            complexity=complexity,
            excel_code=excel_js_code,
            business_description=analysis['business_description'],
            sample_inputs=analysis['sample_inputs'],
            expected_outputs=analysis['expected_outputs'],
            metadata=metadata,
            performance=PerformanceMetrics(
                execution_success_rate=0.85,  # Default for converted models
                user_rating=4.0,
                usage_count=0,
                last_used=None,  # Never used yet
                error_count=0,
                modification_frequency=0.0
            ),
            created_by="xlsx_converter",
            keywords=analysis['keywords'],
            tags=["converted_from_xlsx", f"original_file_{Path(xlsx_path).stem}"]
        )
        
        return model
    
    def _analyze_excel_structure(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Analyze Excel structure to understand the model"""
        analysis = {
            'suggested_name': 'Converted Financial Model',
            'description': 'Financial model converted from Excel',
            'business_description': 'Professional financial model',
            'sample_inputs': {},
            'expected_outputs': {},
            'keywords': [],
            'sections': []
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Look for key sections based on cell content
            for row in sheet.iter_rows(max_row=20, max_col=10, values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        cell_lower = cell.lower()
                        
                        # Detect model type from headers
                        if any(keyword in cell_lower for keyword in ['dcf', 'discounted cash flow']):
                            analysis['keywords'].extend(['dcf', 'valuation', 'enterprise_value'])
                            analysis['suggested_name'] = 'DCF Valuation Model'
                        elif any(keyword in cell_lower for keyword in ['npv', 'net present value']):
                            analysis['keywords'].extend(['npv', 'investment_analysis', 'capital_budgeting'])
                            analysis['suggested_name'] = 'NPV Analysis Model'
                        elif any(keyword in cell_lower for keyword in ['budget', 'forecast']):
                            analysis['keywords'].extend(['budget', 'planning', 'forecast'])
                            analysis['suggested_name'] = 'Budget & Forecast Model'
                        
                        # Detect sections
                        if any(keyword in cell_lower for keyword in ['assumption', 'input']):
                            analysis['sections'].append('assumptions')
                        elif any(keyword in cell_lower for keyword in ['projection', 'forecast']):
                            analysis['sections'].append('projections')
                        elif any(keyword in cell_lower for keyword in ['valuation', 'result', 'summary']):
                            analysis['sections'].append('results')
        
        return analysis
    
    def _generate_excel_js_code(self, workbook: openpyxl.Workbook, analysis: Dict[str, Any]) -> str:
        """Generate Excel.js code from the workbook structure"""
        
        # This is a simplified version - you'd want to make this more sophisticated
        # based on your specific model structures
        
        code_parts = [
            "await Excel.run(async (context) => {",
            "    const sheet = context.workbook.worksheets.getActiveWorksheet();",
            "    ",
            "    // CONVERTED FROM XLSX FILE",
        ]
        
        # Get the main sheet (usually first sheet)
        main_sheet = workbook.worksheets[0]
        
        # Convert key ranges to Excel.js format
        row_count = 0
        for row in main_sheet.iter_rows(max_row=30, max_col=15):
            if row_count > 25:  # Limit for demo
                break
                
            col_values = []
            has_content = False
            
            for cell in row[:10]:  # First 10 columns
                if cell.value is not None:
                    has_content = True
                    if isinstance(cell.value, str):
                        col_values.append(f'"{cell.value}"')
                    elif cell.data_type == 'f':  # Formula
                        # Convert Excel formula to Excel.js format
                        formula = str(cell.value).replace('=', '')
                        col_values.append(f'"={formula}"')
                    else:
                        col_values.append(str(cell.value))
                else:
                    col_values.append('""')
            
            if has_content and col_values:
                row_range = f"A{row_count + 1}:{chr(65 + len(col_values) - 1)}{row_count + 1}"
                values_str = "[" + ", ".join(col_values) + "]"
                code_parts.append(f'    sheet.getRange("{row_range}").values = [{values_str}];')
            
            row_count += 1
        
        code_parts.extend([
            "    ",
            "    // Apply basic formatting",
            '    sheet.getRange("A1").format.font.bold = true;',
            '    sheet.getRange("A1").format.font.size = 14;',
            "    ",
            "    await context.sync();",
            "});"
        ])
        
        return "\n".join(code_parts)
    
    def _extract_metadata(self, workbook: openpyxl.Workbook, analysis: Dict[str, Any]) -> ModelMetadata:
        """Extract metadata from the workbook"""
        
        # Detect Excel functions used
        excel_functions = set()
        components = set(analysis['sections'])
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula
                        formula = str(cell.value).upper()
                        # Extract function names
                        for func in ['NPV', 'IRR', 'SUM', 'AVERAGE', 'VLOOKUP', 'IF', 'PMT', 'FV', 'PV']:
                            if func in formula:
                                excel_functions.add(func)
        
        return ModelMetadata(
            components=list(components) or ['converted_structure'],
            excel_functions=list(excel_functions) or ['SUM'],
            formatting_features=['basic_formatting'],
            business_assumptions=['converted_assumptions'],
            time_horizon_years=5,  # Default
            currencies=['USD'],
            regions=['global']
        )

# Usage example
def convert_excel_collection(excel_directory: str) -> List[FinancialModel]:
    """Convert a directory of Excel files to FinancialModel objects"""
    converter = XLSXToModelConverter()
    models = []
    
    excel_dir = Path(excel_directory)
    
    for excel_file in excel_dir.glob("*.xlsx"):
        try:
            # You'd need to determine these based on filename or content analysis
            model_type = ModelType.DCF  # Could be auto-detected
            industry = Industry.GENERAL  # Could be auto-detected
            complexity = ComplexityLevel.INTERMEDIATE
            
            model = converter.convert_xlsx_to_model(
                str(excel_file),
                f"converted_{excel_file.stem}",
                model_type,
                industry,
                complexity
            )
            
            models.append(model)
            print(f"✅ Converted: {excel_file.name}")
            
        except Exception as e:
            print(f"❌ Failed to convert {excel_file.name}: {e}")
    
    return models

if __name__ == "__main__":
    # Example usage
    models = convert_excel_collection("./financial_models/")
    print(f"Converted {len(models)} models")